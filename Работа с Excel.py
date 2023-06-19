# на входе код получает 2 файла "Doc.xls" и "test.xlsx"
# после работы 1 модуля сохраняется новый файл "test.xls", в котором заполнены ячейки A1 и B1:B7
# после работы 2 модуля сохраняется новый файл "test2.xlsx", в котором заполнены ячейки A1 и B1:B7; C1 и D1:D7

# Модуль 1
import xlrd, xlwt
#открываем файл
rb = xlrd.open_workbook('Doc.xls',formatting_info=True)

#выбираем активный лист
sheet = rb.sheet_by_index(0)

#получаем значение первой ячейки A1
val = sheet.row_values(0)[0]
print(val)

#получаем список значений из всех записей
vals = [sheet.row_values(rownum) for rownum in range(sheet.nrows)]
print(type(vals))
print(vals)

wb = xlwt.Workbook()
ws = wb.add_sheet('Test')
#в A1 записываем значение из ячейки A1 прошлого файла
ws.write(0, 0, val)

#в столбец B запишем нашу последовательность из столбца A исходного файла
i = 0
for rec in vals:
    ws.write(i,1,rec[0])
    i += 1

#сохраняем рабочую книгу
wb.save('test.xls')

# Модуль 2 - тестируем работу с форматом xlsx, для чего делаю копию только созданного файла "test.xls" в новом формате xlsx
import openpyxl
wb = openpyxl.load_workbook(filename = 'test.xlsx')
sheet = wb['Test']

#считываем значение определенной ячейки
val = sheet['A1'].value
print(val)

#считываем заданный диапазон - в указанном примере хоть одна строчка раотает без исправления?
cell_range = [v[0].value for v in sheet['B1:B7']]
print(cell_range)

#записываем значение в определенную ячейку
sheet['C1'] = 'какой же глючный код нам дали'

#записываем последовательность
i = 1
for rec in cell_range:
    sheet.cell(row=i, column=4).value = rec
    i += 1

# сохраняем данные
wb.save('test2.xlsx')

